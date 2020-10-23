import os

import win32api
import xlrd
import xlutils
import xlwt

from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd


def excel(sheet_name):
	# 需求，没有excel时候新建一个，有的时候打开这个excel，新建sheet，sheet里写入内容
	if not os.path.exists("E:\\test.xls"):
		workbook = xlwt.Workbook(encoding='utf-8')  # 新建工作簿
		sheet1 = workbook.add_sheet(sheet_name, 0)  # 新建sheet
		sheet1.write(0, 0, "姓名")  # 第1行第1列数据
		sheet1.write(0, 1, "学号")  # 第1行第2列数据
		sheet1.write(1, 0, "张三")  # 第2行第1列数据
		sheet1.write(1, 1, "036")  # 第2行第2列数据

		workbook.save("E:\\test.xls")  # 保存
		sheet1.write(2, 2, "张三")  # 第2行第1列数据
		sheet1.write(2, 3, "036")  # 第2行第2列数据

		workbook.save("E:\\test.xls")  # 保存
	else:
		# workbook = xlwt.open_workbook("E:\\test.xls")
		# sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
		# sheet1 = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
		# sheet1.write(0, 0, "姓名111")  # 第1行第1列数据
		# sheet1.write(0, 1, "学号111")  # 第1行第2列数据
		# sheet1.write(1, 0, "张三111")  # 第2行第1列数据
		# sheet1.write(1, 1, "036111")  # 第2行第2列数据
		#
		# workbook.save("E:\\test.xls")  # 保存
		# sheet1.write(2, 2, "张三111")  # 第2行第1列数据
		# sheet1.write(2, 3, "036111")  # 第2行第2列数据

		# workbook.save("E:\\test.xls")  # 保存
		oldWb = xlrd.open_workbook("E:\\test.xls")  # 先打开已存在的表
		from xlutils.copy import copy
		newWb = copy(oldWb)  # 复制
		newWs = newWb.get_sheet(0) # 取sheet表
		newWs.write(2, 4, "pass") # 写入 2行4列写入pass
		newWb.save("E:\\test.xls") # 保存至result路径



"""
padas dataframe生成excel
"""
def dataFramesheet(dataframe,excelWriter):
	# DataFrame转换成excel中的sheet表
	dataframe.to_excel(excel_writer=excelWriter, sheet_name="info1",index=None)
	excelWriter.save()
	excelWriter.close()

"""
excel中新增sheet表
"""
def excelAddSheet(dataframe,excelWriter,sheet_name):
	book = load_workbook(excelWriter.path)
	excelWriter.book = book
	dataframe.to_excel(excel_writer=excelWriter,sheet_name=sheet_name,index=None)
	excelWriter.close()


if __name__ == '__main__':
	# excel("111")
	# excel("222")
	# 数据集
	# dataSet = [
	# 	{"姓名": "张三", "年龄": 23, "性别": "男"},
	# 	{"姓名": "李四", "年龄": 25, "性别": "男"},
	# 	{"姓名": "王五", "年龄": 21, "性别": "女"}
	# ]
	dataSet = [
		{"": ""}
	]

	# excelPath
	excelPath = "E:\\test.xlsx"

	# 生成DataFrame
	dataframe = pd.DataFrame(dataSet)

	# 创建ExcelWriter 对象
	excelWriter = pd.ExcelWriter(excelPath, engine='openpyxl')

	# #生成excel
	dataFramesheet(dataframe,excelWriter)

	# excel中增加sheet
	excelAddSheet(dataframe, excelWriter,"11")

	excelAddSheet(dataframe, excelWriter,"22")
	win32api.ShellExecute(0, 'open', excelPath, '', '', 1)
