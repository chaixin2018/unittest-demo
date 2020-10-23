# -*- coding:utf-8 -*-
import time
from time import sleep

import os
import sys
import shutil

import openpyxl
import pyautogui
import selenium
import win32api
import win32con
import win32gui
import xlrd
import xlutils
import xlwt
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from config.global_setting import *

import ctypes

from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd


def clear(self, attr_type, value):
	if attr_type == "xpath":
		self.driver.find_element_by_xpath(value).clear()
	elif attr_type == "id":
		self.driver.find_element_by_id(value).clear()


def wait_result(self, attr_type, value, result):
	if attr_type == "xpath":
		# 超时时间为300秒，每0.2秒检查1次，直到xpath的元素出现
		WebDriverWait(self.driver, 60, 0.2, None).until(EC.text_to_be_present_in_element((By.XPATH, value), result))
		sleep(2)
	elif attr_type == "id":
		WebDriverWait(self.driver, 60, 0.2,None).until(EC.text_to_be_present_in_element((By.ID,value),result))
		sleep(2)


def wait_element(self, attr_type, value):
	if attr_type == "xpath":
		# 超时时间为300秒，每0.2秒检查1次，直到xpath的元素出现
		WebDriverWait(self.driver, 60,0.5, None).until(lambda x: x.find_element_by_xpath(value))
		sleep(2)
	elif attr_type == "id":
		WebDriverWait(self.driver, 60, 0.5, None).until(lambda x: x.find_element_by_id(value))
		sleep(2)


# clickT 自动判断元素是否存在，会有一个延时，但click函数不判断元素是否存在，直接调用，好处是节省时间
def clickT(self, attr_type, value):
	wait_element(self, attr_type, value)
	if attr_type == "xpath":
		self.driver.find_element_by_xpath(value).click()
	elif attr_type == "id":
		self.driver.find_element_by_id(value).click()


def click(self, attr_type, value):
	if attr_type == "xpath":
		self.driver.find_element_by_xpath(value).click()
	elif attr_type == "id":
		self.driver.find_element_by_id(value).click()


def result_action(self, screen_shot_name, result_setting):
	if result_setting == "save_evidence":
		save_evidence(self, screen_shot_name)
	else:
		save_screenshot(self, screen_shot_name)


def save_screenshot1(self, screen_shot_name):
	# 获取报告名称
	report_name = get_report_name()
	# 如果名称是空的，就创建一个，当没有执行suite = unittest.TestSuite()，手动调试的时候就会是空的
	# if report_name == "":
	file = 'TestReport' + time.strftime("%Y%m%d-%H%M", time.localtime())
	os.chdir(report_path)
	if not os.path.exists(report_path):
		# 创建报告文件夹
		os.makedirs(get_path() + '\\Report\\' + file)
		global fileName
		fileName = file
		# 创建函数文件夹
		global file_path
		file_path = get_path() + '\\Report\\' + file + '\\' + get_func_name()
		if not os.path.exists(file_path):
			os.makedirs(file_path)
		global filename
		filename = time.strftime("%H%M%S_", time.localtime()) + screen_shot_name
		self.driver.get_screenshot_as_file(file_path + "\\" + filename + ".png")
	else:
		file_path = get_path() + '\\Report\\' + report_name + '\\' + get_func_name()
		if not os.path.exists(file_path):
			os.makedirs(file_path)
		filename = time.strftime("%H%M%S_", time.localtime()) + screen_shot_name
		self.driver.get_screenshot_as_file(file_path + "\\" + filename + ".png")

def save_evidence1(self, screen_shot_name):

	save_screenshot(self, screen_shot_name)

	# 如果是第一次保留留痕，需要创建excel文件
	evidence_file = file_path + "\\test_evidence.xls"
	if not os.path.exists(evidence_file):

		workbook = xlwt.Workbook(encoding='utf-8')  # 新建工作簿
		sheet1 = workbook.add_sheet("evidence")  # 这里应该是用例名称
		workbook.save(evidence_file)  # 保存
		# 打开截图留痕excel
		win32api.ShellExecute(0, 'open', evidence_file, '', '', 1)
		# 等待excel打开，等待时间是不是有点长
		sleep(10)
		# 获得excel临时文件的句柄
		# handle = win32gui.FindWindow("XLMAIN", "Microsoft Excel - test_evidence.xls  [互換モード]")
		handle = win32gui.FindWindow("XLMAIN", "Microsoft Excel - test_evidence.xls")
		# 激活并显示窗口。
		win32gui.ShowWindow(handle, win32con.SW_MAXIMIZE)
		win32gui.EnableWindow(handle, True)
		win32gui.SetForegroundWindow(handle)
		# 获取最下方缩小excel的图标，并缩小至70%
		sleep(1)
		# 设置输入法为英文
		set_EN()
		pyautogui.moveTo(1311, 848)
		sleep(1)
		pyautogui.click()
		pyautogui.click()
		pyautogui.click()
		sleep(1)
		first_save_evidence = 1

	# 如果是第二次留痕
	else:
		sleep(1)
		# 获得excel临时文件的句柄
		handle = win32gui.FindWindow("XLMAIN", "Microsoft Excel - test_evidence.xls")
		# 激活并显示窗口。
		win32gui.ShowWindow(handle, win32con.SW_MAXIMIZE)
		win32gui.EnableWindow(handle, True)
		win32gui.SetForegroundWindow(handle)
	sleep(2)
	# 点击插入图片，并插入截图好的evidence图片
	pyautogui.keyDown('Alt')
	pyautogui.keyDown("I")
	pyautogui.keyUp('Alt')
	pyautogui.keyUp("I")
	pyautogui.press('P')
	pyautogui.press('F')
	# 鼠标移动到文件路径位置
	time.sleep(2)

	for i in range(0, 6):
		pyautogui.press("Tab")
	sleep(1)
	pyautogui.press("space")
	insert_file_path = file_path
	pyautogui.typewrite(insert_file_path)

	time.sleep(1)
	pyautogui.press('enter')
	for i in range(0, 5):
		pyautogui.press("Tab")
	sleep(1)
	insert_file_name = filename + ".png"
	pyautogui.typewrite(insert_file_name)
	time.sleep(1)
	pyautogui.press('enter')
	sleep(1)

	# 鼠标移动到excel下移图标位置，并移动至空白行，为下一次插入图片的位置做准备
	pyautogui.moveTo(1427, 814)
	for i in range(0, 47):
		pyautogui.click()
	sleep(5)

	# 鼠标移动到excel的初始插入位置
	x, y = 35, 186
	pyautogui.moveTo(x, y)
	pyautogui.click()
	time.sleep(1)

	# 使excel成为disableWindow
	win32gui.ShowWindow(handle, win32con.SW_MINIMIZE)
	win32gui.EnableWindow(handle, False)

# dir_path = ''


def get_base_path():
	# 获取当前的base路径
	global dir_path
	dir_path = os.path.abspath('..')
	global report_path
	report_path = dir_path
	return dir_path


def set_path(path):
	global suite_path
	suite_path = path



def get_path():
	# 获取当前路径
	if dir_path == '':
		get_base_path()

	return dir_path


# fileName = ''
def set_func_name(case_no):
	# 设置用例名称
	global funcName
	funcName = case_no


def get_func_name():
	return funcName


def set_report_name(file):
	global fileName
	fileName = file


def get_report_name():
	return fileName


def close_file(handle,name):
	handle = win32gui.FindWindow(handle,name)
	win32gui.ShowWindow(handle, win32con.SW_MAXIMIZE)
	win32gui.EnableWindow(handle, True)
	win32gui.SetForegroundWindow(handle)
	# 注意这里不可以用Atl+F4
	pyautogui.keyDown('Ctrl')
	pyautogui.keyDown('F4')
	pyautogui.keyUp('Ctrl')
	pyautogui.keyUp("F4")
	sleep(2)
	pyautogui.press('Enter')
	sleep(2)


def getOS_lid_hex():
	sleep(1)
	user32 = ctypes.WinDLL('user32', use_last_error=True)
	curr_window = user32.GetForegroundWindow()
	thread_id = user32.GetWindowThreadProcessId(curr_window, 0)
	klid = user32.GetKeyboardLayout(thread_id)
	lid = klid & (2 ** 16 - 1)
	lid_hex = hex(lid)

	# if lid_hex == '0x409':
	# 	print('当前的输入法状态是英文输入模式\n\n')
	# elif lid_hex == '0x804':
	# 	print('当前的输入法是中文输入模式\n\n')
	# else:
	# 	print('当前的输入法是日语\n\n')
	return lid_hex


def set_EN():
	lid_hex =getOS_lid_hex()
	if lid_hex == '0x409': # 当前是英文输入法
		pass
	elif lid_hex == '0x804': # 当前是中文输入法
		# 按一次ALt+Shift
		pyautogui.keyDown('Alt')
		pyautogui.keyDown('Shift')
		pyautogui.keyUp('Alt')
		pyautogui.keyUp("Shift")
	else: # 当前是日文输入法
		# 按两次ALt+Shift
		pyautogui.keyDown('Alt')
		pyautogui.keyDown('Shift')
		pyautogui.keyUp('Alt')
		pyautogui.keyUp("Shift")
		sleep(1)
		pyautogui.keyDown('Alt')
		pyautogui.keyDown('Shift')
		pyautogui.keyUp('Alt')
		pyautogui.keyUp("Shift")


def set_CN():
	lid_hex =getOS_lid_hex()
	if lid_hex == '0x409': # 当前是英文输入法
		# 按两次ALt+Shift
		pyautogui.keyDown('Alt')
		pyautogui.keyDown('Shift')
		pyautogui.keyUp('Alt')
		pyautogui.keyUp("Shift")
		sleep(1)
		pyautogui.keyDown('Alt')
		pyautogui.keyDown('Shift')
		pyautogui.keyUp('Alt')
		pyautogui.keyUp("Shift")
	elif lid_hex == '0x804': # 当前是中文输入法
		pass
	else: # 当前是日文输入法
		# 按一次ALt+Shift
		pyautogui.keyDown('Alt')
		pyautogui.keyDown('Shift')
		pyautogui.keyUp('Alt')
		pyautogui.keyUp("Shift")


def set_JA():
	lid_hex =getOS_lid_hex()
	if lid_hex == '0x409': # 当前是英文输入法
		# 按一次ALt+Shift
		pyautogui.keyDown('Alt')
		pyautogui.keyDown('Shift')
		pyautogui.keyUp('Alt')
		pyautogui.keyUp("Shift")
	elif lid_hex == '0x804': # 当前是中文输入法
		# 按两次ALt+Shift
		pyautogui.keyDown('Alt')
		pyautogui.keyDown('Shift')
		pyautogui.keyUp('Alt')
		pyautogui.keyUp("Shift")
		sleep(1)
		pyautogui.keyDown('Alt')
		pyautogui.keyDown('Shift')
		pyautogui.keyUp('Alt')
		pyautogui.keyUp("Shift")
	else: # 当前是日文输入法
		pass


def create_report():
	global report_time
	report_time = time.strftime("%Y%m%d-%H%M", time.localtime())
	file = 'TestReport' + report_time
	set_report_name(file)
	path = get_base_path() + '\\Report\\'
	create_name = path + file
	if not os.path.exists(create_name):
		os.makedirs(create_name)
	os.chdir(path + file)
	set_path(path + file)


def create_suite_file(suite_name):
	os.chdir(suite_path)
	path = get_base_path() + '\\'
	file = 'TestReport' + report_time
	create_name = path + file + '\\' + suite_name
	if not os.path.exists(create_name):
		os.makedirs(create_name)
	os.chdir(path + file + '\\' + suite_name)


def set_suite_name(name):
	global suite_name
	suite_name = name


def get_suite_name():
	name = suite_name
	return name


def save_screenshot(self, screen_shot_name):
	create_suite_file(suite_name)
	global screenshot_name
	screenshot_name = time.strftime("%H%M%S_", time.localtime()) + screen_shot_name
	self.driver.get_screenshot_as_file(screenshot_name + ".png")


def save_evidence(self, screen_shot_name):

	save_screenshot(self, screen_shot_name)

	# 如果是第一次保留留痕，需要创建excel文件
	evidence_file = suite_path + "\\" + suite_name + "\\test_evidence.xlsx"
	if not os.path.exists(evidence_file):
		global wb
		wb = Workbook()
		sheet_name = get_func_name()
		ws = wb.create_sheet(sheet_name,0)
		# wb.remove("sheet")
		# ws = wb[sheet_name]
		wb.save(evidence_file)
		wb.close()


		# global workboo
		# workbook = xlwt.Workbook(encoding='utf-8')  # 新建工作簿
		# sheet_name = get_func_name()
		# sheet1 = workbook.add_sheet(sheet_name)  # 这里应该是用例名称
		# # sheet = workbook.add_sheet("evidence")  # 这里应该是用例名称
		# workbook.save(evidence_file)  # 保存

		# 打开截图留痕excel
		win32api.ShellExecute(0, 'open', evidence_file, '', '', 1)
		# 等待excel打开，等待时间是不是有点长
		sleep(10)
		# 获得excel临时文件的句柄
		# handle = win32gui.FindWindow("XLMAIN", "Microsoft Excel - test_evidence.xlsx [互換モード]")
		handle = win32gui.FindWindow("XLMAIN", "Microsoft Excel - test_evidence.xlsx")
		# 激活并显示窗口。
		win32gui.ShowWindow(handle, win32con.SW_MAXIMIZE)
		win32gui.EnableWindow(handle, True)
		win32gui.SetForegroundWindow(handle)
		# 获取最下方缩小excel的图标，并缩小至70%
		sleep(1)
		# 设置输入法为英文
		set_EN()
		pyautogui.moveTo(1311, 848)
		sleep(1)
		pyautogui.click()
		pyautogui.click()
		pyautogui.click()
		sleep(1)
		first_save_evidence = 1

	# 如果是第二次留痕
	else:
		sleep(1)
		# 获得excel临时文件的句柄
		# wb = load_workbook(evidence_file)
		# print("evidence_file",evidence_file)

		# wb = win32api.ShellExecute(0, 'open', evidence_file, '', '', 1)
		# sleep(5)
		# sheet_name = get_func_name()
		# # ws = wb.create_sheet(sheet_name)sleep(10)
		# # 		sheet_name = get_func_name()
		# # wb.save(evidence_file)
		# # print("666")
		# with pd.ExcelWriter(evidence_file) as writer:
		# 	data = "1"
		# 	data.to_excel(writer, sheet_name=sheet_name)

		handle = win32gui.FindWindow("XLMAIN", "Microsoft Excel - test_evidence.xlsx")
		# 激活并显示窗口。
		win32gui.ShowWindow(handle, win32con.SW_MAXIMIZE)
		win32gui.EnableWindow(handle, True)
		win32gui.SetForegroundWindow(handle)


		# sheet_name = get_func_name()
		# print("2workbook", workbook)
		# print("2sheet_name", sheet_name)
		# sleep(2)
		# sheet2 = workbook.add_sheet(sheet_name)  # 这里应该是用例名称
		# # ws1 = workbook.create_sheet("Mysheet")
		# #sheet2 = workbook.add_sheet(sheet_name)  # 这里应该是用例名称
		# #workbook.save(evidence_file)  # 保存
		sleep(2)

	sleep(2)
	# 点击插入图片，并插入截图好的evidence图片
	pyautogui.keyDown('Alt')
	pyautogui.keyDown("I")
	pyautogui.keyUp('Alt')
	pyautogui.keyUp("I")
	pyautogui.press('P')
	pyautogui.press('F')
	# 通过tab选中文件
	time.sleep(2)

	for i in range(0, 6):
		pyautogui.press("Tab")
		sleep(0.5)
	sleep(1)
	pyautogui.press("space")
	insert_file_path = suite_path + "\\" + suite_name
	pyautogui.typewrite(insert_file_path)
	sleep(1)
	pyautogui.press('enter')

	sleep(2)
	for i in range(0, 5):
		pyautogui.press("Tab")
		sleep(0.5)
	sleep(1)
	insert_file_name = screenshot_name + ".png"
	pyautogui.typewrite(insert_file_name)
	time.sleep(1)
	pyautogui.press('enter')
	sleep(1)

	# 鼠标移动到excel下移图标位置，并移动至空白行，为下一次插入图片的位置做准备
	pyautogui.moveTo(1427, 814)
	for i in range(0, 47):
		pyautogui.click()
	sleep(5)

	# 鼠标移动到excel的初始插入位置
	x, y = 35, 186
	pyautogui.moveTo(x, y)
	pyautogui.click()
	time.sleep(1)

	# 使excel成为disableWindow
	win32gui.ShowWindow(handle, win32con.SW_MINIMIZE)
	win32gui.EnableWindow(handle, False)


# 下拉框遍历取值
# 如果option_text是下拉框中内容，就选中这个下拉内容
def is_option_value_present(self,xpath,tag_name,option_text):
	select = self.driver.find_element_by_xpath(xpath)
	# 注意使用find_elements
	options_list = select.find_elements_by_tag_name(tag_name)
	for option in options_list:
		# print ("Value is: " + option.get_attribute("value"))
		# print ("Text is: " +option.text)
		if option.text in option_text:
			select_value = option.get_attribute("value")
			# print ("option_textoption textValue is: " + select_value)
			break

	return select_value


# 遍历下拉框内容
# 如果value_text与下拉框所有遍历后内容一致，则返回选项内容校验正确，否则返回选项内容校验正确
# value_text写法：从第一个依次写，每个选项用， 分割，最后一个后面需要， 举例"製造業, 卸売業, 小売業, IT, "
def option_value_check(self,xpath,tag_name, value_text):
	select = self.driver.find_element_by_xpath(xpath)
	# 注意使用find_elements
	options_list = select.find_elements_by_tag_name(tag_name)
	option_text = ""
	for option in options_list:
		# 获取下拉框的value和text
		# print("Value is:%s  Text is:%s" % (option.get_attribute("value"), option.text))
		s1 = Select(self.driver.find_element_by_xpath(xpath))
		s1.select_by_visible_text(option.text)
		option_text = option_text + option.text  + ', '
		sleep(1)
	if option_text == value_text:
		option_text = option_text + "选项内容校验正确"
	else:
		option_text = option_text + "选项内容校验错误"
	return option_text


# 单选框
# 查找单选按钮并点击
def radio_check(self, xpath, radio_name):
	radios = self.driver.find_elements_by_xpath(xpath)  # find_elements_by_xpath("//*/input[@type='radio']")
	# 注意这里返回的多个单选按钮，返回的是一个列表

	count = 0
	if radios:  # 判断是否有找到元素
		for radio_name in radios:  # 循环点击找到的元素
			radio_name.click()
			count += 1  # 用来记录找到几个单选按钮
			print("打印： ", count)
			time.sleep(2)
	else:
		print("没有找到元素")


# 单选项
# 判断按钮是否被选中
def is_radio_selected(self, xpath, radio_name):
	radios = self.driver.find_elements_by_xpath(xpath)  # 找到页面的单选按钮 "//*/input[@type='radio']"
	count = 0
	for radio in radios:
		is_selected = radio.is_selected()  # 判断按钮是否被选中，选中返回True，没有选中返回false
		if is_selected:  # 返回True时，就执行下面的语句
			count += 1
			print("被选中", " ", count)
		else:
			print("没有被选中", is_selected)


# 复选框
# 查找复选框并点击
def checkboxes_check(self, xpath, check_name):
	checkboxes = self.driver.find_elements_by_xpath("xpath")
	count = 0
	if checkboxes:  # 判断是否有找到元素

		for check_name in checkboxes:  # 循环点击找到的元素
			check_name.click()  # 勾选复选框
			count += 1
			print("打印信息 ", count)
			time.sleep(2)
	else:
		print("没有找到元素")


# 文本框
def textboxes_check(self, xpath, max, min, text_type, text_name):
	textoxes = self.driver.find_elements_by_xpath("//*[@id='phone-form']//input[@type='checkbox']")
	count = 0
	if textoxes:  # 判断是否有找到元素

		for text_name in textoxes:  # 循环点击找到的元素
			text_name.click()  # 勾选复选框
			count += 1
			print("打印信息 ", count)
			time.sleep(2)
	else:
		print("没有找到元素")


if __name__ == '__main__':
	pass